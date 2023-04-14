from openpyxl import load_workbook
import gspread

wb = load_workbook('CCM.xlsx')
ws = wb['кэш']
gc = gspread.service_account(filename='pidor-of-the-day-af3dd140b860.json')  # доступ к гугл табл по ключевому файлу аккаунта разраба
sh = gc.open('CCM')
worksheet2 = sh.worksheet('заявки')
list_of_lists = worksheet2.get_all_values()
list_of_dicts = worksheet2.get_all_records()
#print(list_of_dicts[1]['идентификатор'])
cell_id = {}
name = []
status = []
# cell_id[list_of_dicts[0]['номер заказа']] = float(list_of_dicts[0]['цена итого'].replace('\xa0', '').replace(',', '.'))
# print(cell_id)
for i in list_of_dicts:
    if i[' telegram id'] is not None:
        try:
            if i[' telegram id'] == 127154290 and i['статус заказа (оформил клиент заказ или же он просто в корзине)'] == 'TRUE':
                cell_id[str(i['номер заказа'])] = {'name': f'{cell_id[str(i["номер заказа"])]["name"]}\n'
                                                           f'{i["товар"]} - {i["количество"]} шт.',
                                                   'price': cell_id[str(i["номер заказа"])]["price"] +
                                                   float(i['цена итого'].replace('\xa0', '').replace(',', '.')),
                                                   'status': f'{cell_id[str(i["номер заказа"])]["status"]}'}
        except KeyError:
            cell_id[str(i['номер заказа'])] = {'name': f'{i["товар"]} - {i["количество"]} шт.', 'price':
                                               float(i['цена итого'].replace('\xa0', '').replace(',', '.')), 'status':
                                               f'{i["статус заказа (отображается у клиента)"]}'}
            status = i['статус заказа (отображается у клиента)']
    else:
        break


for i in reversed(cell_id.keys()):
    name.append(f'Заказ № {i}\n'
                f'{cell_id[i]["name"]}\n'
                f'Cтатус: {cell_id[i]["status"]}\n'
                f'К оплате: {cell_id[i]["price"]}\n\n')
name = ' '.join(name)
print(name)

# def zakazy_search(self):
#     name = []
#     zakazy = []
#     price = 0
#     status = []
#     list_of_lists = worksheet2.get_all_values()
#     bot.send_message(message.chat.id, "Собираем данные..")
#
#     cell_id = (self.worksheet2.findall(str(self.message.chat.id), in_column=1))[::-1]
#     for i in cell_id:
#         if self.worksheet2.cell(i.row, 15).value == 'TRUE':
#             zakazy.append(self.worksheet2.cell(i.row, 13).value)
#     if len(set(zakazy)) != 0:
#         for a in set(zakazy):
#             name.append(f' Заказ № {a}\n')
#             for i in cell_id:
#                 if self.worksheet2.cell(i.row, 13).value == a:
#                     name.append(
#                         f'{self.worksheet2.cell(i.row, 6).value} - {self.worksheet2.cell(i.row, 7).value} шт.\n')
#                     price += float(self.worksheet2.cell(i.row, 11).value.replace(",", "."))
#                     status = self.worksheet2.cell(i.row, 16).value
#             name.append(f'Статус: {status}\n К оплате: {price} ₽\n\n')
#         name = ' '.join(name)
#         self.bot.send_message(self.message.chat.id, f' Ваша история заказов:\n'
#                                                     f'{name}')
#     else:
#         self.bot.send_message(self.message.chat.id, f'Заказы отсутствуют')