import gspread  # библиотека работы с гугл таблицами
from telebot import types
from datetime import *  # библиотека проверки даты
from passwords import *
from openpyxl import load_workbook  # библиотека работы с exel таблицами


tovar_descriptions = None
ostatok = None
admin_id = igor


class buttons:  # класс для создания клавиатур различных категорий товаров
    global file, tovar_row

    def __init__(self, bot, message, kategoriya=None, list_one=None,
                 image='https://drive.google.com/file/d/1nG0RvJ9L6Ez_O9SOjllhFn2OvszB92TE/view?usp=share_link'):
        self.bot = bot
        self.message = message
        self.image = image
        self.kategoriya = kategoriya
        self.list_one = list_one

    def menu_buttons(self):
        kb1 = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)
        but1 = types.KeyboardButton(text='Каталог 🗂️')
        but2 = types.KeyboardButton(text='Поступления 🆕')
        but3 = types.KeyboardButton(text='Распродажа 🏷️')
        but4 = types.KeyboardButton(text='О нас ⁉️')
        but5 = types.KeyboardButton(text='Мои заказы 📋')
        but6 = types.KeyboardButton(text='Корзина 🗑️')
        kb1.add(but1, but2, but3, but4, but5, but6)
        self.bot.send_message(self.message.chat.id, text='...', reply_markup=kb1)

    def razdely_buttons(self):
        keys = {}
        kb1 = types.InlineKeyboardMarkup()
        for i in self.list_one:
            keys[f'but{self.list_one.index(i)}'] = types.InlineKeyboardButton(text=i, callback_data=i)
            if self.list_one.index(i) > 0 and self.list_one.index(i) % 2 != 0:
                if len(i) <= 16 and len(self.list_one[self.list_one.index(i) - 1]) <= 16:
                    kb1.add(keys[f'but{self.list_one.index(i) - 1}'], keys[f'but{self.list_one.index(i)}'])  # , row_width=1)
                else:
                    kb1.add(keys[f'but{self.list_one.index(i) - 1}'])
                    kb1.add(keys[f'but{self.list_one.index(i)}'])
            elif self.list_one.index(i) == (len(self.list_one) - 1):
                kb1.add(keys[f'but{self.list_one.index(i)}'])
        self.bot.send_photo(self.message.chat.id, photo=self.image)
        self.bot.send_message(self.message.chat.id, text=f'Пожалуйста выберите {self.kategoriya}', reply_markup=kb1)

    def marks_buttons(self):  # функция создающая клавиатуру
        keys = {}
        kb1 = types.InlineKeyboardMarkup()
        for i in self.list_one:
            keys[f'but{self.list_one.index(i)}'] = types.InlineKeyboardButton(text=i[0], callback_data=i[1])
            if self.list_one.index(i) > 0 and self.list_one.index(i) % 2 != 0:
                if len(i[0]) <= 16 and len(self.list_one[self.list_one.index(i) - 1][0]) <= 16:
                    kb1.add(keys[f'but{self.list_one.index(i) - 1}'],
                            keys[f'but{self.list_one.index(i)}'])  # , row_width=1)
                else:
                    try:
                        kb1.add(keys[f'but{self.list_one.index(i) - 1}'])
                        kb1.add(keys[f'but{self.list_one.index(i)}'])
                    except Exception:
                        self.bot.send_message(admin_id, 'Ошибка!!!Проверьте таблицу на дублирование')
            elif self.list_one.index(i) == (len(self.list_one) - 1):
                kb1.add(keys[f'but{self.list_one.index(i)}'])
        self.bot.send_photo(self.message.chat.id, photo=self.image)
        self.bot.send_message(self.message.chat.id, text=f'Пожалуйста выберите {self.kategoriya}', reply_markup=kb1)

    def zayavka_buttons(self, back_value='Вернуться в начало'):
        kb4 = types.InlineKeyboardMarkup(row_width=1)
        but1 = types.InlineKeyboardButton(text='Добавить товар в корзину!', callback_data='Да, хочу!')
        but3 = types.InlineKeyboardButton(text='Вернуться назад', callback_data=back_value)
        kb4.add(but1, but3)
        self.bot.send_message(self.message.chat.id, f'Хотите оформить заявку/купить онлайн выбранный товар?\n '
                                                    f'/help - Подробности покупки', reply_markup=kb4)

    def basket_buttons(self):
        kb7 = types.InlineKeyboardMarkup(row_width=1)
        but1 = types.InlineKeyboardButton(text='Оформить заказ', callback_data='Оформить заказ')
        but2 = types.InlineKeyboardButton(text='Редактировать корзину', callback_data='redact')
        but3 = types.InlineKeyboardButton(text='Очистить корзину', callback_data='delete_row')
        kb7.add(but1, but2, but3)
        self.bot.send_message(self.message.chat.id, f'Хотите оформить заказ/купить онлайн выбранный товар?\n '
                                                    f'/help - Подробности покупки', reply_markup=kb7)

    def basket_buttons_redact(self):
        keys = {}
        gc = gspread.service_account(filename='pidor-of-the-day-af3dd140b860.json')  # доступ к гугл табл по ключевому
        # файлу аккаунта разраба
        try:
            sh = gc.open('CCM')
            worksheet2 = sh.worksheet('заявки')
            cell_id = (worksheet2.findall(str(self.message.chat.id), in_column=1))[::-1]
            kb4 = types.InlineKeyboardMarkup()
            rows = []
            for i in cell_id:
                if worksheet2.cell(i.row, 15).value == 'FALSE':
                    keys[f'but{cell_id.index(i)}'] = types.InlineKeyboardButton(text=int(cell_id.index(i)) + 1,
                                                                                callback_data=f'red_row{i.row}')
                    kb4.add(keys[f'but{cell_id.index(i)}'])
                    rows.append(i.row)
            keys[f'but{len(cell_id) + 1}'] = types.InlineKeyboardButton(text='Вернуться назад', callback_data=f'Корзина')
            kb4.add(keys[f'but{len(cell_id) + 1}'])
            self.bot.edit_message_text('Выберите номер позиции в которой нужно изменить количество',
                                       self.message.chat.id, self.message.id)
            self.bot.edit_message_reply_markup(self.message.chat.id, self.message.id, reply_markup=kb4)
        except Exception:
            self.bot.send_message(self.message.chat.id, 'Ошибка подключения. Повторите запрос через 1 минуту.')


def tovar_in_basket(bot, message):
    global ostatok
    wb = load_workbook('CCM.xlsx')
    ws = wb['кэш']
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=10, values_only=True):
        if row[0] == message.chat.id:
            quantity = row[5]
            bot.send_message(message.chat.id,
                             f'Товар успешно добавлен в корзину.\n Воспользуйтесь основным меню чтобы '
                             f'продолжить покупки.🤝\n')
            buttons(bot, message).menu_buttons()
            poisk_tovar_in_base(bot, message, row[2], row[1], quantity, size=row[3],
                                price=row[4], dostavka=row[8]).zayavka_v_baze(
                                                             int(quantity)*(float(str(row[4]).replace(",", ".").replace(
                                                                 " ", "").replace("\xa0", ""))), row[9])
            break


def zayavka_done(bot, message, number):
    gc = gspread.service_account(filename='pidor-of-the-day-af3dd140b860.json')  # доступ к гугл табл по ключевому файлу аккаунта разраба
    try:
        sh = gc.open('CCM')
        worksheet2 = sh.worksheet('заявки')
        cell_id = (worksheet2.findall(str(message.chat.id), in_column=1))[::-1]# поиск ячейки с данными по ключевому слову
        zakaz_id = f'{message.chat.id}{cell_id[0].row}'
        for i in cell_id:
            if worksheet2.cell(i.row, 15).value == 'FALSE':
                worksheet2.update_cell(i.row, 5, number)
                worksheet2.update_cell(i.row, 13, zakaz_id)
                worksheet2.update_cell(i.row, 15, 'TRUE')
        bot.send_message(message.chat.id,
                         f'Вашему заказу присвоен номер {zakaz_id}. Заказ отправлен менеджеру.\n'
                         f'Дождитесь обратного звонка в ближайшее время, для уточнения деталей заказа.')
        buttons(bot, message).menu_buttons()
        bot.send_message(admin_id, f'🚨!!!ВНИМАНИЕ!!!🚨\n'
                                   f'Поступила ЗАЯВКА от:\n'
                                   f'id чата: {message.chat.id}\n'
                                   f'Имя: {message.chat.first_name}\n'
                                   f'Фамилия: {message.chat.last_name}\n'
                                   f'№ телефона: {number}\n'
                                   f'Ссылка: @{message.chat.username}\n'
                                   f'Номер заказа: {zakaz_id}')
        bot.send_message(admin_id, 'Заявка внесена в базу ✅\n'
                                   'смотреть базу: https://docs.google.com/spreadsheets/d/'
                                   '14P5j3t4Z9kmy4o87WEbLqeTwsKi7YZAx7RiQPlY2c1w/edit?usp=sharing')

    except Exception:
        bot.send_message(message.chat.id, 'Ошибка подключения. Повторите запрос через 1 минуту.')
        buttons(bot, message).menu_buttons()


class poisk_tovar_in_base:

    def __init__(self, bot, message, article=None, tovar_name=None, vnalichii=None, image=None, size=None,
                 price=None, your_price=None, size_web=None, tovar_type=None, dostavka=None):
        self.bot = bot
        self.message = message
        self.article = article
        self.tovar_name = tovar_name
        self.vnalichii = vnalichii
        self.image = image
        self.size = size
        self.price = price
        self.your_price = your_price
        self.size_web = size_web
        self.tovar_type = tovar_type
        self.dostavka = dostavka
        self.wb = load_workbook('CCM.xlsx')
        self.ws = self.wb['кэш']
        self.ws2 = self.wb['МЛ Остатки штаб']
        gc = gspread.service_account(
            filename='pidor-of-the-day-af3dd140b860.json')  # доступ к гугл табл по ключевому файлу аккаунта разраба
        # открытие таблицы по юрл адресу:
        try:
            sh = gc.open('CCM')
            self.worksheet2 = sh.worksheet('заявки')
        except Exception:
            self.bot.send_message(self.message.chat.id, 'Ошибка подключения. Повторите запрос через 1 минуту.')

    def poisk_ostatok(self, back_value='Вернуться в начало'):
        global ostatok
        try:
            self.bot.send_message(self.message.chat.id, 'Проверяем наличие..')
            if self.image is not None:
                self.bot.send_photo(self.message.chat.id, self.image, f'{self.tovar_name}\nРазмер: {self.size}\n'
                                                                      f'Артикул: {self.article}')
                self.bot.send_message(self.message.chat.id, f'В наличии: {self.vnalichii}\n'
                                                            f'Тип товара: {self.tovar_type}\nПрайс: {self.price} ₽\n'
                                                            f'Цена опт: {self.your_price} ₽\nДоставка: {self.dostavka}\n'
                                                            f'Таблица размеров: {self.size_web}')
                if self.ws.max_row >= 20:
                    self.ws.delete_rows(5, self.ws.max_row)
                    self.ws.insert_rows(1)
                    self.ws['A1'] = self.message.chat.id
                    self.ws['B1'] = self.tovar_name
                    self.ws['C1'] = self.article
                    self.ws['D1'] = self.size
                    self.ws['E1'] = self.your_price
                    self.ws['G1'] = self.vnalichii
                    self.ws['I1'] = self.dostavka
                    self.ws['J1'] = back_value
                    self.bot.send_message(self.message.chat.id, 'загружаем базу данных..')
                    self.ws['H1'] = poisk_tovar_in_base(self.bot, self.message).poisk_number()
                    self.wb.save('CCM.xlsx')
                else:
                    self.ws.insert_rows(1)
                    self.ws['A1'] = self.message.chat.id
                    self.ws['B1'] = self.tovar_name
                    self.ws['C1'] = self.article
                    self.ws['D1'] = self.size
                    self.ws['E1'] = self.your_price
                    self.ws['G1'] = self.vnalichii
                    self.ws['I1'] = self.dostavka
                    self.ws['J1'] = back_value
                    self.bot.send_message(self.message.chat.id, 'загружаем базу данных..')
                    self.ws['H1'] = poisk_tovar_in_base(self.bot, self.message).poisk_number()
                    self.wb.save('CCM.xlsx')
                if self.vnalichii == 0:
                    kb4 = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
                    but1 = types.KeyboardButton(text='Вернуться в начало')
                    kb4.add(but1)
                    self.bot.send_message(self.message.chat.id, f'Увы товар закончился\n'
                                                                f'/help - справка по боту\n', reply_markup=kb4)
                else:
                    buttons(self.bot, self.message).zayavka_buttons(back_value=back_value)
                    ostatok = self.vnalichii
            else:
                self.bot.send_message(self.message.chat.id, f'Фото товара временно отсутствует\n{self.tovar_name}\n'
                                                            f'Размер: {self.size}\nАртикул: {self.article}')
                self.bot.send_message(self.message.chat.id, f'В наличии: {self.vnalichii}\n'
                                                            f'Тип товара: {self.tovar_type}\nПрайс: {self.price}\n'
                                                            f'Цена опт: {self.your_price}\nДоставка: {self.dostavka}\n'
                                                            f'Таблица размеров: {self.size_web}')
                if self.ws.max_row >= 20:
                    self.ws.delete_rows(5, self.ws.max_row)
                    self.ws.insert_rows(1)
                    self.ws['A1'] = self.message.chat.id
                    self.ws['B1'] = self.tovar_name
                    self.ws['C1'] = self.article
                    self.ws['D1'] = self.size
                    self.ws['E1'] = self.your_price
                    self.ws['G1'] = self.vnalichii
                    self.ws['I1'] = self.dostavka
                    self.ws['J1'] = back_value
                    self.bot.send_message(self.message.chat.id, 'загружаем базу данных..')
                    self.ws['H1'] = poisk_tovar_in_base(self.bot, self.message).poisk_number()
                    self.wb.save('CCM.xlsx')
                else:
                    self.ws.insert_rows(1)
                    self.ws['A1'] = self.message.chat.id
                    self.ws['B1'] = self.tovar_name
                    self.ws['C1'] = self.article
                    self.ws['D1'] = self.size
                    self.ws['E1'] = self.your_price
                    self.ws['G1'] = self.vnalichii
                    self.ws['I1'] = self.dostavka
                    self.ws['J1'] = back_value
                    self.bot.send_message(self.message.chat.id, 'загружаем базу данных..')
                    self.ws['H1'] = poisk_tovar_in_base(self.bot, self.message).poisk_number()
                    self.wb.save('CCM.xlsx')
                if self.vnalichii == 0:
                    kb4 = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
                    but1 = types.KeyboardButton(text='Вернуться в начало')
                    kb4.add(but1)
                    self.bot.send_message(self.message.chat.id, f'Увы товар закончился\n'
                                                                f'/help - справка по боту\n', reply_markup=kb4)
                else:
                    buttons(self.bot, self.message).zayavka_buttons(back_value=back_value)
                    ostatok = self.vnalichii
        except Exception:
            self.bot.send_message(self.message.chat.id, f'Ошибка, товар отсутствует')

    def zayavka_v_baze(self, itogo, tovar_kategory):  # функция перевода из базы потенциальных клиентов в базу старых клиентов
        cell_id = (self.worksheet2.findall(str(self.message.chat.id), in_column=1))[::-1] # поиск ячейки с данными по ключевому слову
        try:
            for i in cell_id:
                if self.worksheet2.cell(i.row, 15).value == 'FALSE' and str(self.worksheet2.cell(i.row, 9).value) == str(self.article):
                    self.worksheet2.update(f'G{i.row}:L{i.row}',
                                           [[int(self.worksheet2.cell(i.row, 7).value) + int(self.vnalichii),
                                             str(datetime.now().date()), self.worksheet2.cell(i.row, 9).value,
                                             self.worksheet2.cell(i.row, 10).value,
                                             (int(self.worksheet2.cell(i.row, 7).value) + int(self.vnalichii)) *
                                             float(str(self.worksheet2.cell(i.row, 10).value).replace(",", ".").replace(" ", "").replace("\xa0", "")),
                                             self.dostavka]])
                    for a in range(1, self.ws2.max_row + 1):
                        if str(self.ws2.cell(a, 1).value) == str(self.article):
                            self.ws2.cell(a, 8).value = int(self.ws2.cell(a, 8).value) - int(self.vnalichii)
                            self.wb.save('CCM.xlsx')
                            break
                    break
            else:
                worksheet_len2 = len(self.worksheet2.col_values(1)) + 1  # запись клиента в свободную строку базы старых клиентов:
                self.worksheet2.update(f'A{worksheet_len2}:N{worksheet_len2}',
                                       [[self.message.chat.id, self.message.from_user.username,
                                         self.message.from_user.first_name, self.message.from_user.last_name, None,
                                         self.tovar_name, self.vnalichii, str(datetime.now().date()), self.article,
                                         self.price, itogo, self.dostavka, None, tovar_kategory]])
                for a in range(1, self.ws2.max_row + 1):
                    if str(self.ws2.cell(a, 1).value) == str(self.article):
                        self.ws2.cell(a, 8).value = int(self.ws2.cell(a, 8).value) - int(self.vnalichii)
                        self.wb.save('CCM.xlsx')
                        break
        except AttributeError:
            self.bot.send_message(admin_id, 'Ошибка добавления  товара в базу')

    def poisk_number(self):
        cell_id = self.worksheet2.find(str(self.message.chat.id), in_column=1)
        if cell_id is not None:
            number = self.worksheet2.cell(cell_id.row, 5).value
            return number
        else:
            return None

    def basket_search(self):
        name = []
        sum_price = 0
        sum_quantity = 0
        self.bot.send_message(self.message.chat.id, "Собираем данные..")
        cell_id = (self.worksheet2.findall(str(self.message.chat.id), in_column=1))[::-1]
        for i in cell_id:
            if self.worksheet2.cell(i.row, 15).value == 'FALSE':
                name.append(f'\nПозиция: {cell_id.index(i) + 1}\n'
                            f'{self.worksheet2.cell(i.row, 6).value} - {self.worksheet2.cell(i.row, 7).value} шт. \n'
                            f'Категория - {self.worksheet2.cell(i.row, 14).value}\n'
                            f'Цена - {self.worksheet2.cell(i.row, 11).value}₽\n'
                            f'Доставка - {self.worksheet2.cell(i.row, 12).value}\n')
                sum_price += float(str(self.worksheet2.cell(i.row, 11).value).replace(",", ".").replace(" ", "").replace("\xa0", ""))
                sum_quantity += int(self.worksheet2.cell(i.row, 7).value)
        name_ = ' '.join(name)
        if len(name) != 0:
            self.bot.send_message(self.message.chat.id, f'На данный момент Ваши товары в корзине:\n'
                                                        f'{name_}\n'
                                                        f'Итого в корзине товаров: \n'
                                                        f'в количестве {sum_quantity} шт.\n'
                                                        f'на сумму {sum_price}₽')
            buttons(self.bot, self.message).basket_buttons()
        else:
            self.bot.send_message(self.message.chat.id, f'Товары отсутствуют..Проверьте историю заказов.')

    def basket_delete(self, row):
        row = int(row)
        try:
            article = self.worksheet2.cell(row, 9).value
            for a in range(1, self.ws2.max_row + 1):
                if str(self.ws2.cell(a, 1).value) == str(article):
                    self.ws2.cell(a, 8).value = int(self.ws2.cell(a, 8).value) + int(self.worksheet2.cell(row, 7).value)
                    self.bot.send_message(self.message.chat.id, 'Товар успешно удален из корзины')
                    self.worksheet2.batch_clear([f"A{row}:N{row}"])
                    self.wb.save('CCM.xlsx')
                    break
        except AttributeError:
            self.bot.send_message(self.message.chat.id, 'Товар уже был удален ранее. Перейдите в корзину снова, чтобы '
                                                        'обновить данные')

    def basket_delete_all(self):
        try:
            cell_id = (self.worksheet2.findall(str(self.message.chat.id), in_column=1))[::-1]
            for i in cell_id:
                if self.worksheet2.cell(i.row, 15).value == 'FALSE':
                    article = self.worksheet2.cell(i.row, 9).value
                    for a in range(1, self.ws2.max_row + 1):
                        if str(self.ws2.cell(a, 1).value) == str(article):
                            self.ws2.cell(a, 8).value = int(self.ws2.cell(a, 8).value) + int(self.worksheet2.cell(i.row, 7).value)
                            self.worksheet2.batch_clear([f"A{i.row}:N{i.row}"])
                            self.wb.save('CCM.xlsx')
                            break
            self.bot.send_message(self.message.chat.id, 'Товары успешно удалены')
        except AttributeError:
            self.bot.send_message(self.message.chat.id, 'Товар уже был удален ранее. Перейдите в корзину снова, чтобы '
                                                        'обновить данные')

    def zakazy_search(self):
        list_of_dicts = self.worksheet2.get_all_records()
        cell_id = {}
        for i in list_of_dicts:
            if i[' telegram id'] is not None:
                try:
                    if i[' telegram id'] == self.message.chat.id and i[
                         'статус заказа (оформил клиент заказ или же он просто в корзине)'] == 'TRUE':
                        cell_id[str(i['номер заказа'])] = {'name': f'{cell_id[str(i["номер заказа"])]["name"]}\n'
                                                                   f'{i["товар"]} - {i["количество"]} шт.',
                                                           'price': cell_id[str(i["номер заказа"])]["price"] + float(
                                                                    i['цена итого'].replace('\xa0', '').replace(',', '.')),
                                                           'status': f'{cell_id[str(i["номер заказа"])]["status"]}'}
                except KeyError:
                    cell_id[str(i['номер заказа'])] = {'name': f'{i["товар"]} - {i["количество"]} шт.', 'price':
                                                       float(i['цена итого'].replace('\xa0', '').replace(',', '.')),
                                                       'status': f'{i["статус заказа (отображается у клиента)"]}'}
            else:
                break
        if len(cell_id.keys()) != 0:
            block = []
            for i in reversed(cell_id.keys()):
                block.append(f'Заказ № {i}\n'
                             f'{cell_id[i]["name"]}\n'
                             f'Cтатус: {cell_id[i]["status"]}\n'
                             f'К оплате: {cell_id[i]["price"]} ₽\n\n')
            block = ' '.join(block)
            self.bot.send_message(self.message.chat.id, f' Ваша история заказов:\n'
                                                        f'{block}')
        else:
            self.bot.send_message(self.message.chat.id, f'Заказы отсутствуют')


class rasylka_message:  # класс хранения сообщения для рассылки
    def __init__(self, post):
        self.post = post

    def _get_message_(self):
        return self.post
