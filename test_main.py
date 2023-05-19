import telebot  # библиотека телеграм-бота
from telebot import types  # с помощью типов можно создавать клавиатуры
import gspread
from openpyxl import load_workbook
from apscheduler.schedulers.background import BackgroundScheduler

from test_functions import buttons, poisk_tovar_in_base, rasylka_message, admin_id, tovar_in_basket, zayavka_done, \
    statistic
from passwords import *
article = None

token = code_mashine
# token = lemonade
# token = ccmclub
bot = telebot.TeleBot(token)

tovar_name = None
quantity = None
rassylka = None


@bot.message_handler(commands=['start'])    # перехватчик команды /start
def start(message):
    file_open = open("start_logo.png", 'rb')    # открытие и чтение файла стартовой картинки
    bot.send_photo(message.chat.id, file_open, '''Здравствуйте!
Вас приветствует бот CCM_Club.
Я помогу подобрать хоккейный инвентарь из наличия по лучшим ценам. 🏆🏒🥇

Выберите «Каталог 🗂️» для просмотра товаров.
«О нас» расскажет вам о нас и как мы работаем.
Команда в строке /help – о всех возможностях бота.
''')
    buttons(bot, message).menu_buttons()


@bot.message_handler(commands=['help'])
def help(message):
    kb1 = types.InlineKeyboardMarkup()
    with open('command_help.txt', 'r') as help_text:
        help_text1 = help_text.read()
    if message.chat.id == admin_id:      # условия демонстрации различных команд для админа и клиентов
        but1 = types.InlineKeyboardButton('редактировать текст команды /help', callback_data='rhelp')
        kb1.add(but1)
        bot.send_message(message.chat.id, help_text1,
                         reply_markup=kb1)
    else:
        bot.send_message(message.chat.id, help_text1)
        buttons(bot, message).menu_buttons()


@bot.message_handler(commands=['sent_message'])
def sent_message(message):
    if message.chat.id == admin_id:
        sent = bot.send_message(admin_id, 'Введи id чата клиента, которому нужно написать от лица бота')
        bot.register_next_step_handler(sent, sent_message_perehvat_1)   # перехватывает ответ пользователя на сообщение "sent" и
                                                              # и направляет его аргументом в функцию base_perehvat
    else:
        bot.send_message(message.chat.id, 'У Вас нет прав для использования данной команды')


@bot.message_handler(func=lambda m: m.text)  # перехватчик текстовых сообщений
def chek_message_category(m):
    list_one = []
    back_value = 'Вернуться в начало'
    wb = load_workbook('CCM.xlsx')
    ws = wb['МЛ Остатки штаб']
    if m.text == 'Каталог 🗂️':
        for row in ws.iter_rows(min_row=2, min_col=9, max_col=9, values_only=True):
            if row == (None,):
                break
            list_one.append(*row)
        list_one = sorted(list(set(list_one)))
        file_open = open("menu_logo.jpeg", 'rb')
        buttons(bot, m, kategoriya='раздел', list_one=list_one,
                image=file_open).razdely_buttons()
        statistic().proverka(m)
    elif m.text == 'Поступления 🆕':
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=15, values_only=True):
            if row[14] is not None:
                list_one.append(f'🆕{row[8]}')
            elif row[8] == (None,):
                break
        list_one = sorted(list(set(list_one)))
        file_open = open("menu_logo.jpeg", 'rb')
        buttons(bot, m, kategoriya='раздел', list_one=list_one,
                image=file_open).razdely_buttons()
        statistic().proverka(m)
    elif m.text == 'Распродажа 🏷️':
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=15, values_only=True):
            if row[9] is not None:
                list_one.append(f'🏷️{row[8]}')
            elif row[8] == (None,):
                break
        list_one = sorted(list(set(list_one)))
        file_open = open("menu_logo.jpeg", 'rb')
        buttons(bot, m, kategoriya='раздел', list_one=list_one,
                image=file_open).razdely_buttons()
        statistic().proverka(m)
    elif m.text == "Вернуться в корзину":
        bot.send_message(m.chat.id, f'Загружаем..')
        poisk_tovar_in_base(bot, m).basket_search()
        buttons(bot, m).menu_buttons()
    elif m.text == 'Мои заказы 📋':
        bot.send_message(m.chat.id, 'Загружаем..')
        poisk_tovar_in_base(bot, m).zakazy_search()
    elif m.text == 'Корзина 🗑️':
        bot.send_message(m.chat.id, f'Загружаем..')
        poisk_tovar_in_base(bot, m).basket_search()
    elif m.text == 'О нас ⁉️':
        kb1 = types.InlineKeyboardMarkup()
        with open('about.txt', 'r') as help_text:
            help_text1 = help_text.read()
        if m.chat.id == admin_id:  # условия демонстрации различных команд для админа и клиентов
            but1 = types.InlineKeyboardButton('редактировать текст раздела "О нас"', callback_data='rabout')
            kb1.add(but1)
            bot.send_message(m.chat.id, help_text1,
                             reply_markup=kb1)
        else:
            bot.send_message(m.chat.id, help_text1)
    elif len(list_one) == 0:
        list_two = []
        kategoriya = None
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=16, values_only=True):
            if row == (None,):
                break
            elif row[8] == m.text:     # если колбек равен разделу
                list_one.append(row[1][0:30])
                list_one = sorted(list(set(list_one)))
                kategoriya = 'категорию'
                back_value = 'Вернуться в начало'
            elif m.text in str(row[1]):
                if len(row[2]) <= 25:
                    list_two.append((str(row[2])+'-'+str(row[3]), row[0]))
                    back_value = row[8]
                    kategoriya = 'товар'
                else:
                    list_two.append((f'{row[2][:15]}...{str((row[2])+str(row[3]))[-12:]}', row[0]))
                    back_value = row[8]
                    kategoriya = 'товар'
            elif m.text in str(row[2]):
                tovar_name = row[2]
                article = row[0]
                image = row[10]
                size = row[3]
                price = row[4]
                vnalichii = row[7]
                tovar_type = row[15]
                your_price = row[5]
                dostavka = row[11]
                size_web = row[13]
                bot.send_message(m.chat.id, 'Загружаем..')
                poisk_tovar_in_base(bot, m, article, vnalichii=vnalichii, tovar_name=tovar_name,
                                    image=image, size=size, price=price,
                                    your_price=your_price, size_web=size_web, tovar_type=tovar_type,
                                    dostavka=dostavka).poisk_ostatok(back_value=row[1])
        if len(list_one) != 0:
            file_open = open("menu_logo.jpeg", 'rb')
            list_one.append(back_value)
            buttons(bot, m, kategoriya=kategoriya, list_one=list_one,
                    image=file_open).razdely_buttons()
        elif len(list_two) != 0:
            file_open = open("menu_logo.jpeg", 'rb')
            list_two.append(('Вернуться назад', back_value))
            buttons(bot, m, kategoriya=kategoriya, list_one=list_two,
                    image=file_open).marks_buttons()


@bot.callback_query_handler(func=lambda callback: callback.data)
def check_callback(callback):
    global tovar_name, quantity, article
    back_value = 'Вернуться в начало'
    wb = load_workbook('CCM.xlsx')
    ws = wb['МЛ Остатки штаб']
    list_one = []
    if callback.data == 'Да, хочу!':
        val = bot.send_message(callback.message.chat.id,
                               'Пожалуйста отправьте количество желаемого товара ЧИСЛОМ с помощью клавиатуры')
        bot.register_next_step_handler(val, amount)
    elif callback.data == 'Оформить заказ':
        redact_basket(bot, callback.message).zapros_number()
    elif callback.data[:7] == 'red_row':
        val = bot.edit_message_text(f'Пожалуйста введите НОВОЕ значение количества товара ЧИСЛОМ с помощью клавиатуры.\n'
                                    f'Для удаления позиции введите ноль (0).',
                                    callback.message.chat.id, callback.message.id)
        bot.register_next_step_handler(val, redact_basket(bot, callback, callback.data[7:]).redact_quintity)
    elif callback.data == 'delete_row':
        bot.send_message(callback.message.chat.id, f'Подчищаем базу..')
        poisk_tovar_in_base(bot, callback.message).basket_delete_all()
    elif callback.data == 'rhelp':
        kb7 = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        but1 = types.KeyboardButton(text='Каталог 🗂️')
        kb7.add(but1)
        val = bot.send_message(callback.message.chat.id, 'введите новый текст команды /help', reply_markup=kb7)
        bot.register_next_step_handler(val, redact_basket(bot, callback.message, file='command_help.txt').redact_text)
    elif callback.data == 'rabout':
        kb7 = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        but1 = types.KeyboardButton(text='Каталог 🗂️')
        kb7.add(but1)
        val = bot.send_message(callback.message.chat.id, 'введите новый текст раздела "О нас"', reply_markup=kb7)
        bot.register_next_step_handler(val, redact_basket(bot, callback.message, file='about.txt').redact_text)
    elif callback.data == 'Корзина':
        kb7 = types.InlineKeyboardMarkup(row_width=1)
        but1 = types.InlineKeyboardButton(text='Оформить заказ', callback_data='Оформить заказ')
        but2 = types.InlineKeyboardButton(text='Редактировать корзину', callback_data="redact")
        but3 = types.InlineKeyboardButton(text='Очистить корзину', callback_data='delete_row')
        kb7.add(but1, but2, but3)
        bot.edit_message_text(f'Хотите оформить заказ/купить онлайн выбранный товар?\n '
                              f'/help - Подробности покупки', callback.message.chat.id,
                              callback.message.id)
        bot.edit_message_reply_markup(callback.message.chat.id, callback.message.id, reply_markup=kb7)
    elif callback.data == 'redact':
        buttons(bot, callback.message).basket_buttons_redact()
    elif callback.data == "Вернуться в начало":   # кнопка "вернуться в начало" для каталога
        for row in ws.iter_rows(min_row=2, min_col=9,   max_col=9, values_only=True):
            if row == (None,):
                break
            list_one.append(*row)
        list_one = list(set(list_one))
        file_open = open("menu_logo.jpeg", 'rb')
        buttons(bot, callback.message, kategoriya='раздел', list_one=list_one,
                image=file_open).razdely_buttons()
    elif callback.data == '🆕Вернуться в начало':     # кнопка "вернуться в начало" для поступлений
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=15, values_only=True):
            if row[14] is not None:
                list_one.append(f'🆕{row[8]}')
            elif row[8] == (None,):
                break
        list_one = list(set(list_one))
        file_open = open("menu_logo.jpeg", 'rb')
        buttons(bot, callback.message, kategoriya='раздел', list_one=list_one,
                image=file_open).razdely_buttons()
    elif callback.data == '🏷️Вернуться в начало':     # кнопка "вернуться в начало" для поступлений
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=15, values_only=True):
            if row[9] is not None:
                list_one.append(f'🏷️{row[8]}')
            elif row[8] == (None,):
                break
        list_one = list(set(list_one))
        file_open = open("menu_logo.jpeg", 'rb')
        buttons(bot, callback.message, kategoriya='раздел', list_one=list_one,
                image=file_open).razdely_buttons()

    elif len(list_one) == 0:
        list_two = []
        kategoriya = None
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=16, values_only=True):
            if row == (None,):
                break
            elif f'🆕{row[8]}' == callback.data:  # если колбек равен разделу из поступлений
                if row[14] is not None:
                    list_one.append(f'🆕{row[1][0:30]}')
                    list_one = sorted(list(set(list_one)))
                    kategoriya = 'категорию'
                    back_value = '🆕Вернуться в начало'
            elif row[8] == callback.data:     # если колбек равен разделу
                list_one.append(row[1][0:30])
                list_one = sorted(list(set(list_one)))
                kategoriya = 'категорию'
                back_value = 'Вернуться в начало'
            elif f'🏷️{row[8]}' == callback.data:  # если колбек равен разделу из поступлений
                if row[9] is not None:
                    list_one.append(f'🏷️{row[1][0:30]}')
                    list_one = sorted(list(set(list_one)))
                    kategoriya = 'категорию'
                    back_value = '🏷️Вернуться в начало'
            elif (callback.data in str(f'🆕{row[1]}')) and ('🆕' in callback.data) and row[14] is not None:  # если колбек содержится в категории из поступлений, содержит смайл нев и колонка поступление не пустая
                if len(row[2]) <= 25:  # проверка длины названия категории
                    list_two.append((f'🆕{str(row[2])}' + '-' + str(row[3]), f'🆕{row[0]}')) # формирование наименования из имени и размера, второе значение - атрибут
                    back_value = f'🆕{row[8]}'  # метка для кнопки вернуться назад (будет возвращаться в разделы поступлений)
                    kategoriya = 'товар'
                else:
                    list_two.append((f'🆕{row[2][:15]}...{str((row[2]) + str(row[3]))[-12:]}', f'🆕{row[0]}'))
                    back_value = f'🆕{row[8]}'
                    kategoriya = 'товар'
            elif (callback.data in str(f'🏷️{row[1]}')) and ('🏷️' in callback.data) and row[9] is not None:  # если колбек содержится в категории из поступлений, содержит смайл нев и колонка поступление не пустая
                if len(row[2]) <= 25:  # проверка длины названия категории
                    list_two.append((f'🏷️{str(row[2])}' + '-' + str(row[3]), f'🏷️{row[0]}')) # формирование наименования из имени и размера, второе значение - атрибут
                    back_value = f'🏷️{row[8]}'  # метка для кнопки вернуться назад (будет возвращаться в разделы поступлений)
                    kategoriya = 'товар'
                else:
                    list_two.append((f'🏷️{row[2][:15]}...{str((row[2]) + str(row[3]))[-12:]}', f'🏷️{row[0]}'))
                    back_value = f'🏷️{row[8]}'
                    kategoriya = 'товар'
            elif callback.data in str(row[1]):
                if len(row[2]) <= 25:
                    list_two.append((str(row[2])+'-'+str(row[3]), row[0]))
                    back_value = row[8]
                    kategoriya = 'товар'
                else:
                    list_two.append((f'{row[2][:15]}...{str((row[2])+str(row[3]))[-12:]}', row[0]))
                    back_value = row[8]
                    kategoriya = 'товар'
            elif str(row[0]) == str(callback.data):
                tovar_name = row[2]
                article = row[0]
                image = row[10]
                size = row[3]
                price = row[4]
                vnalichii = row[7]
                tovar_type = row[15]
                your_price = row[5]
                dostavka = row[11]
                size_web = row[13]
                bot.send_message(callback.message.chat.id, 'Загружаем..')
                poisk_tovar_in_base(bot, callback.message, article, vnalichii=vnalichii, tovar_name=tovar_name,
                                    image=image, size=size, price=price,
                                    your_price=your_price, size_web=size_web, tovar_type=tovar_type,
                                    dostavka=dostavka).poisk_ostatok(back_value=row[1])
            elif f'🆕{row[0]}' == str(callback.data):
                tovar_name = row[2]
                article = row[0]
                image = row[10]
                size = row[3]
                price = row[4]
                vnalichii = row[7]
                tovar_type = row[15]
                your_price = row[5]
                dostavka = row[11]
                size_web = row[13]
                bot.send_message(callback.message.chat.id, 'Загружаем..')
                poisk_tovar_in_base(bot, callback.message, article, vnalichii=vnalichii, tovar_name=tovar_name,
                                    image=image, size=size, price=price,
                                    your_price=your_price, size_web=size_web, tovar_type=tovar_type,
                                    dostavka=dostavka).poisk_ostatok(back_value=f'🆕{row[1]}')
            elif f'🏷️{row[0]}' == str(callback.data):
                tovar_name = row[2]
                article = row[0]
                image = row[10]
                size = row[3]
                price = row[4]
                vnalichii = row[7]
                tovar_type = row[15]
                your_price = row[5]
                dostavka = row[11]
                size_web = row[13]
                bot.send_message(callback.message.chat.id, 'Загружаем..')
                poisk_tovar_in_base(bot, callback.message, article, vnalichii=vnalichii, tovar_name=tovar_name,
                                    image=image, size=size, price=price,
                                    your_price=your_price, size_web=size_web, tovar_type=tovar_type,
                                    dostavka=dostavka).poisk_ostatok(back_value=f'🏷️{row[1]}')
        if len(list_one) != 0:
            file_open = open("menu_logo.jpeg", 'rb')
            list_one.append(back_value)
            buttons(bot, callback.message, kategoriya=kategoriya, list_one=list_one,
                    image=file_open).razdely_buttons()
        elif len(list_two) != 0:
            file_open = open("menu_logo.jpeg", 'rb')
            list_two.append(('Вернуться назад', back_value))
            buttons(bot, callback.message, kategoriya=kategoriya, list_one=list_two,
                    image=file_open).marks_buttons()


def amount(message):
    wb = load_workbook('CCM.xlsx')
    ws = wb['кэш']
    try:
        int(message.text)
        for i in range(1, ws.max_row + 1):
            if str(ws.cell(i, 1).value) == str(message.chat.id):
                if int(message.text) <= int(ws.cell(i, 7).value) and int(message.text) != 0:
                    ws.cell(i, 6).value = message.text
                    wb.save('CCM.xlsx')
                    tovar_in_basket(bot=bot, message=message)
                    break
                else:
                    bot.send_message(message.chat.id,
                                     f'Увы, но указанное количество либо превышает остатки товара, либо равно 0. '
                                     f'Отправьте корректное значение.\n'
                                     f'Чтобы изменить товар выберите "Категории товаров 🗂️"')
                    buttons(bot, message).zayavka_buttons()
                    break
    except ValueError:
        bot.send_message(message.chat.id, f'Пожалуйста, укажите количество ЧИСЛОМ')
        buttons(bot, message).zayavka_buttons()


def sent_message_perehvat_1(message):
    try:
        global rasylka
        rasylka = rasylka_message(message.text)  # хз почему message.id а не message.text но bot.copy_message() работает только так
        sent = bot.send_message(admin_id, 'Введите текст сообщения')
        bot.register_next_step_handler(sent, sent_message_perehvat_2)
    except ValueError:
        bot.send_message(admin_id, 'Неккоректное значение. Воспользуйтесь командой /sent_message еще раз')


def sent_message_perehvat_2(message):
    kb2 = types.ReplyKeyboardRemove()
    global rasylka
    bot.copy_message(rasylka.post, admin_id, message.id, reply_markup=kb2)
    bot.send_message(admin_id, 'Сообщение отправлено!')


class redact_basket:
    def __init__(self, bot, message, i=None, file=None):
        self.bot = bot
        self.message = message
        self.i = i
        self.file = file
        gc = gspread.service_account(
            filename='pidor-of-the-day-af3dd140b860.json')  # доступ к гугл табл по ключевому файлу аккаунта разраба
        # открытие таблицы по юрл адресу:
        try:
            sh = gc.open('CCM')
            self.worksheet2 = sh.worksheet('заявки')
        except Exception:
            self.bot.send_message(self.message.chat.id, 'Ошибка подключения. Повторите запрос через 1 минуту.')

    def redact_quintity(self, message):
        wb = load_workbook('CCM.xlsx')
        ws2 = wb['МЛ Остатки штаб']
        try:
            int(message.text)
            if int(message.text) == 0:
                bot.send_message(message.chat.id, f'Подчищаем базу..')
                poisk_tovar_in_base(bot, message).basket_delete(self.i)
                bot.send_message(message.chat.id, f'Загружаем..')
                poisk_tovar_in_base(bot, message).basket_search()
            else:
                row = int(self.i)
                article = self.worksheet2.cell(row, 9).value
                for a in range(1, ws2.max_row + 1):
                    if str(ws2.cell(a, 1).value) == str(article):
                        if (int(ws2.cell(a, 8).value) + int(self.worksheet2.cell(row, 7).value)) >= int(message.text):
                            ws2.cell(a, 8).value = int(ws2.cell(a, 8).value) + int(self.worksheet2.cell(row, 7).value) - \
                                                   int(message.text)
                            bot.send_message(message.chat.id, 'Данные в корзине изменены. Благодарим за ваш выбор!')
                            self.worksheet2.update_cell(row, 7, int(message.text))
                            self.worksheet2.update_cell(row, 11, int(self.worksheet2.cell(row, 7).value) *
                                                        float(str(self.worksheet2.cell(row, 10).value).replace(",", ".").replace(" ", "").replace("\xa0", "")))
                            poisk_tovar_in_base(bot, message).basket_search()
                            wb.save('CCM.xlsx')
                            break
                        else:
                            self.bot.edit_message_text(f'Увы, но указанное количество превышает остатки товара. '
                                                       f'Отправьте корректное значение.', self.message.message.chat.id,
                                                       self.message.message.id)
                            buttons(self.bot, self.message.message).basket_buttons_redact()
                            break
        except ValueError:
            bot.send_message(message.chat.id, f'Пожалуйста, укажите количество ЧИСЛОМ')
            buttons(bot, message).zayavka_buttons()

    def zapros_number(self):
        wb = load_workbook('CCM.xlsx')
        ws = wb['кэш']
        for i in range(1, ws.max_row + 1):
            if str(ws.cell(i, 1).value) == str(self.message.chat.id):
                if ws.cell(i, 8).value is not None:
                    zayavka_done(bot=self.bot, message=self.message, number=ws.cell(i, 8).value)
                    break
        else:
            kb4 = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            but1 = types.KeyboardButton(text='Вернуться в корзину')
            kb4.add(but1)
            mes_num = bot.send_message(self.message.chat.id, 'Для оформления заказа и передачи его менеджеру просим '
                                                             'ввести ваш контактный номер (с помощью клавиатуры).',
                                       reply_markup=kb4)
            bot.register_next_step_handler(mes_num, save_number)

    def redact_text(self, message):
        if message.text == 'Каталог 🗂️':
            chek_message_category(message)
            buttons(self.bot, message).menu_buttons()
        else:
            with open(self.file, 'w') as help_txt:
                help_txt.write(message.text)
                buttons(self.bot, message).menu_buttons()
                bot.send_message(message.chat.id, 'текст успешно изменен')


def save_number(message):
    number = message.text
    if number == "Вернуться в корзину":
        chek_message_category(message)
    elif len(str(number)) >= 10 and ('+' in (str(number)) or '7' in (str(number)) or '8' in (str(number))):
        zayavka_done(bot=bot, message=message, number=number)
    else:
        bot.send_message(message.chat.id, 'указан некорректный номер.')
        kb4 = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
        but1 = types.KeyboardButton(text='Вернуться в начало')
        kb4.add(but1)
        mes_num = bot.send_message(message.chat.id, 'Для оформления заказа и передачи его менеджеру просим '
                                                    'ввести ваш контактный номер (с помощью клавиатуры).',
                                   reply_markup=kb4)
        bot.register_next_step_handler(mes_num, save_number)


scheduler = BackgroundScheduler()
scheduler.add_job(statistic().obnulenie, "cron", day_of_week='mon-sun', hour=0)
#scheduler.add_job(statistic().obnulenie, "interval", seconds=5)
scheduler.start()

bot.infinity_polling()
